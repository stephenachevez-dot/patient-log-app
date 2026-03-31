
import base64
import hashlib
import hmac
import io
import secrets
import sqlite3
import tempfile
from collections import defaultdict
from contextlib import closing
from datetime import date, datetime
from pathlib import Path
from typing import Optional, Tuple

import matplotlib.pyplot as plt
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas

APP_TITLE = "Care Family Health Center | Patient Log App"
DB_PATH = Path(__file__).parent / "patient_log.db"


def get_conn() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    return conn


CONN = get_conn()


def init_db() -> None:
    with closing(CONN.cursor()) as cur:
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                full_name TEXT NOT NULL,
                email TEXT NOT NULL UNIQUE,
                password_hash TEXT NOT NULL,
                role TEXT NOT NULL CHECK(role IN ('provider', 'patient')),
                dob TEXT,
                language TEXT DEFAULT 'English',
                provider_id INTEGER,
                created_at TEXT NOT NULL,
                FOREIGN KEY(provider_id) REFERENCES users(id)
            )
            """
        )
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS glucose_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                patient_user_id INTEGER NOT NULL,
                entered_by_user_id INTEGER NOT NULL,
                log_date TEXT NOT NULL,
                fasting_bs REAL,
                breakfast_notes TEXT,
                before_lunch_bs REAL,
                lunch_notes TEXT,
                before_dinner_bs REAL,
                dinner_notes TEXT,
                before_bedtime_bs REAL,
                nighttime_snack_notes TEXT,
                comments TEXT,
                created_at TEXT NOT NULL,
                FOREIGN KEY(patient_user_id) REFERENCES users(id),
                FOREIGN KEY(entered_by_user_id) REFERENCES users(id)
            )
            """
        )
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS vitals_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                patient_user_id INTEGER NOT NULL,
                entered_by_user_id INTEGER NOT NULL,
                log_date TEXT NOT NULL,
                time_period TEXT,
                blood_pressure TEXT,
                weight_lbs REAL,
                event_notes TEXT,
                created_at TEXT NOT NULL,
                FOREIGN KEY(patient_user_id) REFERENCES users(id),
                FOREIGN KEY(entered_by_user_id) REFERENCES users(id)
            )
            """
        )
    CONN.commit()


def hash_password(password: str, salt: Optional[bytes] = None) -> str:
    salt = salt or secrets.token_bytes(16)
    digest = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt, 200_000)
    return base64.b64encode(salt + digest).decode("utf-8")


def verify_password(password: str, stored_value: str) -> bool:
    raw = base64.b64decode(stored_value.encode("utf-8"))
    salt, saved_digest = raw[:16], raw[16:]
    check_digest = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt, 200_000)
    return hmac.compare_digest(saved_digest, check_digest)


def get_user_by_email(email: str):
    return CONN.execute(
        "SELECT * FROM users WHERE lower(email) = lower(?)",
        (email.strip(),),
    ).fetchone()


def get_user_by_id(user_id: int):
    return CONN.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()


def email_exists(email: str) -> bool:
    return get_user_by_email(email) is not None


def create_user(full_name: str, email: str, password: str, role: str, dob: Optional[str], language: str) -> None:
    CONN.execute(
        """
        INSERT INTO users (full_name, email, password_hash, role, dob, language, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?)
        """,
        (
            full_name.strip(),
            email.strip().lower(),
            hash_password(password),
            role,
            dob,
            language,
            datetime.utcnow().isoformat(),
        ),
    )
    CONN.commit()


def assign_patient_to_provider(patient_user_id: int, provider_user_id: int) -> None:
    CONN.execute(
        "UPDATE users SET provider_id = ? WHERE id = ? AND role = 'patient'",
        (provider_user_id, patient_user_id),
    )
    CONN.commit()


def get_provider_patients(provider_user_id: int, search_text: str = ""):
    like = f"%{search_text.strip().lower()}%"
    return CONN.execute(
        """
        SELECT * FROM users
        WHERE role = 'patient'
          AND provider_id = ?
          AND (
            ? = '%%'
            OR lower(full_name) LIKE ?
            OR lower(email) LIKE ?
          )
        ORDER BY full_name
        """,
        (provider_user_id, like, like, like),
    ).fetchall()


def get_all_unassigned_patients():
    return CONN.execute(
        """
        SELECT * FROM users
        WHERE role = 'patient' AND provider_id IS NULL
        ORDER BY full_name
        """
    ).fetchall()


def delete_patient_account(patient_user_id: int, provider_user_id: int) -> bool:
    patient = CONN.execute(
        "SELECT * FROM users WHERE id = ? AND role = 'patient' AND provider_id = ?",
        (patient_user_id, provider_user_id),
    ).fetchone()
    if not patient:
        return False

    with closing(CONN.cursor()) as cur:
        cur.execute("DELETE FROM glucose_logs WHERE patient_user_id = ?", (patient_user_id,))
        cur.execute("DELETE FROM vitals_logs WHERE patient_user_id = ?", (patient_user_id,))
        cur.execute("DELETE FROM users WHERE id = ?", (patient_user_id,))
    CONN.commit()
    return True


def add_glucose_log(
    patient_user_id: int,
    entered_by_user_id: int,
    log_date: str,
    fasting_bs,
    breakfast_notes,
    before_lunch_bs,
    lunch_notes,
    before_dinner_bs,
    dinner_notes,
    before_bedtime_bs,
    nighttime_snack_notes,
    comments,
) -> None:
    CONN.execute(
        """
        INSERT INTO glucose_logs (
            patient_user_id, entered_by_user_id, log_date,
            fasting_bs, breakfast_notes, before_lunch_bs, lunch_notes,
            before_dinner_bs, dinner_notes, before_bedtime_bs,
            nighttime_snack_notes, comments, created_at
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            patient_user_id,
            entered_by_user_id,
            log_date,
            fasting_bs,
            breakfast_notes,
            before_lunch_bs,
            lunch_notes,
            before_dinner_bs,
            dinner_notes,
            before_bedtime_bs,
            nighttime_snack_notes,
            comments,
            datetime.utcnow().isoformat(),
        ),
    )
    CONN.commit()


def update_glucose_log(
    log_id: int,
    patient_user_id: int,
    log_date: str,
    fasting_bs,
    breakfast_notes,
    before_lunch_bs,
    lunch_notes,
    before_dinner_bs,
    dinner_notes,
    before_bedtime_bs,
    nighttime_snack_notes,
    comments,
) -> None:
    CONN.execute(
        """
        UPDATE glucose_logs
        SET log_date = ?, fasting_bs = ?, breakfast_notes = ?, before_lunch_bs = ?,
            lunch_notes = ?, before_dinner_bs = ?, dinner_notes = ?, before_bedtime_bs = ?,
            nighttime_snack_notes = ?, comments = ?
        WHERE id = ? AND patient_user_id = ?
        """,
        (
            log_date,
            fasting_bs,
            breakfast_notes,
            before_lunch_bs,
            lunch_notes,
            before_dinner_bs,
            dinner_notes,
            before_bedtime_bs,
            nighttime_snack_notes,
            comments,
            log_id,
            patient_user_id,
        ),
    )
    CONN.commit()


def delete_glucose_log(log_id: int, patient_user_id: int) -> None:
    CONN.execute("DELETE FROM glucose_logs WHERE id = ? AND patient_user_id = ?", (log_id, patient_user_id))
    CONN.commit()


def add_vitals_log(
    patient_user_id: int,
    entered_by_user_id: int,
    log_date: str,
    time_period: str,
    blood_pressure: str,
    weight_lbs,
    event_notes: str,
) -> None:
    CONN.execute(
        """
        INSERT INTO vitals_logs (
            patient_user_id, entered_by_user_id, log_date,
            time_period, blood_pressure, weight_lbs, event_notes, created_at
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            patient_user_id,
            entered_by_user_id,
            log_date,
            time_period,
            blood_pressure,
            weight_lbs,
            event_notes,
            datetime.utcnow().isoformat(),
        ),
    )
    CONN.commit()


def update_vitals_log(
    log_id: int,
    patient_user_id: int,
    log_date: str,
    time_period: str,
    blood_pressure: str,
    weight_lbs,
    event_notes: str,
) -> None:
    CONN.execute(
        """
        UPDATE vitals_logs
        SET log_date = ?, time_period = ?, blood_pressure = ?, weight_lbs = ?, event_notes = ?
        WHERE id = ? AND patient_user_id = ?
        """,
        (log_date, time_period, blood_pressure, weight_lbs, event_notes, log_id, patient_user_id),
    )
    CONN.commit()


def delete_vitals_log(log_id: int, patient_user_id: int) -> None:
    CONN.execute("DELETE FROM vitals_logs WHERE id = ? AND patient_user_id = ?", (log_id, patient_user_id))
    CONN.commit()


def get_glucose_logs(patient_user_id: int):
    return CONN.execute(
        "SELECT * FROM glucose_logs WHERE patient_user_id = ? ORDER BY log_date DESC, id DESC",
        (patient_user_id,),
    ).fetchall()


def get_vitals_logs(patient_user_id: int):
    return CONN.execute(
        "SELECT * FROM vitals_logs WHERE patient_user_id = ? ORDER BY log_date DESC, id DESC",
        (patient_user_id,),
    ).fetchall()


def get_recent_glucose_logs(patient_user_id: int, limit: int = 5):
    return CONN.execute(
        "SELECT * FROM glucose_logs WHERE patient_user_id = ? ORDER BY log_date DESC, id DESC LIMIT ?",
        (patient_user_id, limit),
    ).fetchall()


def get_vitals_logs_grouped_by_day(patient_user_id: int):
    rows = get_vitals_logs(patient_user_id)
    grouped = defaultdict(list)
    for row in rows:
        grouped[row["log_date"]].append(row)
    return dict(grouped)


def validate_bp(bp: str) -> bool:
    if not bp:
        return True
    parts = bp.split("/")
    if len(parts) != 2:
        return False
    return all(p.strip().isdigit() for p in parts)


def parse_bp(bp: str) -> Tuple[Optional[int], Optional[int]]:
    if not bp or not validate_bp(bp):
        return None, None
    systolic, diastolic = bp.split("/")
    return int(systolic.strip()), int(diastolic.strip())


def glucose_chart_df(patient_user_id: int) -> pd.DataFrame:
    rows = get_glucose_logs(patient_user_id)
    records = []
    for row in rows:
        records.append(
            {
                "log_date": row["log_date"],
                "fasting_bs": row["fasting_bs"],
                "before_lunch_bs": row["before_lunch_bs"],
                "before_dinner_bs": row["before_dinner_bs"],
                "before_bedtime_bs": row["before_bedtime_bs"],
            }
        )
    df = pd.DataFrame(records)
    if df.empty:
        return df
    df["log_date"] = pd.to_datetime(df["log_date"])
    return df.sort_values("log_date")


def vitals_chart_df(patient_user_id: int) -> pd.DataFrame:
    rows = get_vitals_logs(patient_user_id)
    records = []
    for row in rows:
        systolic, diastolic = parse_bp(row["blood_pressure"] or "")
        records.append(
            {
                "log_date": row["log_date"],
                "time_period": row["time_period"],
                "blood_pressure": row["blood_pressure"],
                "systolic": systolic,
                "diastolic": diastolic,
                "weight_lbs": row["weight_lbs"],
            }
        )
    df = pd.DataFrame(records)
    if df.empty:
        return df
    df["log_date"] = pd.to_datetime(df["log_date"])
    return df.sort_values(["log_date", "time_period"])


def glucose_flag(value):
    if value is None:
        return ("No reading", "info")
    if value < 70:
        return ("Low", "error")
    if value > 250:
        return ("Very High", "error")
    if value > 180:
        return ("High", "warning")
    return ("In Range", "success")


def bp_flag(bp: str):
    systolic, diastolic = parse_bp(bp or "")
    if systolic is None or diastolic is None:
        return ("No reading", "info")
    if systolic >= 180 or diastolic >= 120:
        return ("Urgent", "error")
    if systolic >= 140 or diastolic >= 90:
        return ("High", "warning")
    return ("In Range", "success")


def weight_change_flag(current_weight, previous_weight):
    if current_weight in (None, "") or previous_weight in (None, ""):
        return None
    try:
        diff = float(current_weight) - float(previous_weight)
    except Exception:
        return None
    if abs(diff) >= 5:
        direction = "gain" if diff > 0 else "loss"
        return (f"Rapid {direction} ({diff:+.1f} lbs)", "warning")
    return None


def render_flag(label: str, level: str):
    if level == "error":
        st.error(label)
    elif level == "warning":
        st.warning(label)
    elif level == "success":
        st.success(label)
    else:
        st.info(label)


def show_alert_summary(patient_user_id: int):
    glucose_rows = get_glucose_logs(patient_user_id)
    vitals_rows = get_vitals_logs(patient_user_id)

    low_glucose = 0
    high_glucose = 0
    urgent_bp = 0
    high_bp = 0

    for row in glucose_rows:
        for key in ["fasting_bs", "before_lunch_bs", "before_dinner_bs", "before_bedtime_bs"]:
            value = row[key]
            if value is None:
                continue
            if value < 70:
                low_glucose += 1
            elif value > 180:
                high_glucose += 1

    for row in vitals_rows:
        systolic, diastolic = parse_bp(row["blood_pressure"] or "")
        if systolic is None or diastolic is None:
            continue
        if systolic >= 180 or diastolic >= 120:
            urgent_bp += 1
        elif systolic >= 140 or diastolic >= 90:
            high_bp += 1

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Low Glucose Flags", low_glucose)
    c2.metric("High Glucose Flags", high_glucose)
    c3.metric("High BP Flags", high_bp)
    c4.metric("Urgent BP Flags", urgent_bp)


def apply_clinical_chart_style(ax, title: str, ylabel: str):
    ax.set_title(title, fontsize=12, pad=12, fontweight="bold")
    ax.set_xlabel("Date")
    ax.set_ylabel(ylabel)
    ax.grid(True, alpha=0.3, linestyle="--")
    ax.tick_params(axis="x", rotation=45)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)


def show_trend_charts(patient_user_id: int):
    st.subheader("Trend Charts")

    glucose_df = glucose_chart_df(patient_user_id)
    vitals_df = vitals_chart_df(patient_user_id)

    if glucose_df.empty and vitals_df.empty:
        st.info("Add entries to see trend charts.")
        return

    if not glucose_df.empty:
        fig1, ax1 = plt.subplots(figsize=(8, 4))
        plotted = False
        for col, label in [
            ("fasting_bs", "Fasting"),
            ("before_lunch_bs", "Before Lunch"),
            ("before_dinner_bs", "Before Dinner"),
            ("before_bedtime_bs", "Before Bedtime"),
        ]:
            series = glucose_df[["log_date", col]].dropna()
            if not series.empty:
                ax1.plot(series["log_date"], series[col], marker="o", linewidth=2, label=label)
                plotted = True
        ax1.axhline(70, linestyle="--", linewidth=1)
        ax1.axhline(180, linestyle="--", linewidth=1)
        ax1.text(glucose_df["log_date"].min(), 72, "Low threshold", fontsize=8)
        ax1.text(glucose_df["log_date"].min(), 182, "High threshold", fontsize=8)
        apply_clinical_chart_style(ax1, "Blood Sugar Trend", "mg/dL")
        if plotted:
            ax1.legend(frameon=False)
        fig1.tight_layout()
        st.pyplot(fig1)

    if not vitals_df.empty:
        bp_df = vitals_df.dropna(subset=["systolic", "diastolic"]).copy()
        if not bp_df.empty:
            fig2, ax2 = plt.subplots(figsize=(8, 4))
            ax2.plot(bp_df["log_date"], bp_df["systolic"], marker="o", linewidth=2, label="Systolic")
            ax2.plot(bp_df["log_date"], bp_df["diastolic"], marker="o", linewidth=2, label="Diastolic")
            ax2.axhline(130, linestyle="--", linewidth=1)
            ax2.axhline(80, linestyle="--", linewidth=1)
            ax2.text(bp_df["log_date"].min(), 132, "Systolic ref", fontsize=8)
            ax2.text(bp_df["log_date"].min(), 82, "Diastolic ref", fontsize=8)
            apply_clinical_chart_style(ax2, "Blood Pressure Trend", "mmHg")
            ax2.legend(frameon=False)
            fig2.tight_layout()
            st.pyplot(fig2)

        weight_df = vitals_df.dropna(subset=["weight_lbs"]).copy()
        if not weight_df.empty:
            fig3, ax3 = plt.subplots(figsize=(8, 4))
            ax3.plot(weight_df["log_date"], weight_df["weight_lbs"], marker="o", linewidth=2)
            apply_clinical_chart_style(ax3, "Weight Trend", "lbs")
            fig3.tight_layout()
            st.pyplot(fig3)


def save_chart_image(fig) -> str:
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".png")
    fig.savefig(tmp.name, bbox_inches="tight", dpi=180)
    plt.close(fig)
    return tmp.name


def build_chart_images(patient_user_id: int):
    images = []

    glucose_df = glucose_chart_df(patient_user_id)
    vitals_df = vitals_chart_df(patient_user_id)

    if not glucose_df.empty:
        fig1, ax1 = plt.subplots(figsize=(7, 3.5))
        plotted = False
        for col, label in [
            ("fasting_bs", "Fasting"),
            ("before_lunch_bs", "Before Lunch"),
            ("before_dinner_bs", "Before Dinner"),
            ("before_bedtime_bs", "Before Bedtime"),
        ]:
            series = glucose_df[["log_date", col]].dropna()
            if not series.empty:
                ax1.plot(series["log_date"], series[col], marker="o", linewidth=2, label=label)
                plotted = True
        ax1.axhline(70, linestyle="--", linewidth=1)
        ax1.axhline(180, linestyle="--", linewidth=1)
        ax1.text(glucose_df["log_date"].min(), 72, "Low threshold", fontsize=8)
        ax1.text(glucose_df["log_date"].min(), 182, "High threshold", fontsize=8)
        apply_clinical_chart_style(ax1, "Blood Sugar Trend", "mg/dL")
        if plotted:
            ax1.legend(frameon=False)
        fig1.tight_layout()
        images.append(("Blood Sugar Trend", save_chart_image(fig1)))

    if not vitals_df.empty:
        bp_df = vitals_df.dropna(subset=["systolic", "diastolic"]).copy()
        if not bp_df.empty:
            fig2, ax2 = plt.subplots(figsize=(7, 3.5))
            ax2.plot(bp_df["log_date"], bp_df["systolic"], marker="o", linewidth=2, label="Systolic")
            ax2.plot(bp_df["log_date"], bp_df["diastolic"], marker="o", linewidth=2, label="Diastolic")
            ax2.axhline(130, linestyle="--", linewidth=1)
            ax2.axhline(80, linestyle="--", linewidth=1)
            ax2.text(bp_df["log_date"].min(), 132, "Systolic ref", fontsize=8)
            ax2.text(bp_df["log_date"].min(), 82, "Diastolic ref", fontsize=8)
            apply_clinical_chart_style(ax2, "Blood Pressure Trend", "mmHg")
            ax2.legend(frameon=False)
            fig2.tight_layout()
            images.append(("Blood Pressure Trend", save_chart_image(fig2)))

        weight_df = vitals_df.dropna(subset=["weight_lbs"]).copy()
        if not weight_df.empty:
            fig3, ax3 = plt.subplots(figsize=(7, 3.5))
            ax3.plot(weight_df["log_date"], weight_df["weight_lbs"], marker="o", linewidth=2)
            apply_clinical_chart_style(ax3, "Weight Trend", "lbs")
            fig3.tight_layout()
            images.append(("Weight Trend", save_chart_image(fig3)))

    return images


def build_patient_excel(patient_user_id: int) -> bytes:
    patient = get_user_by_id(patient_user_id)
    glucose_logs = get_glucose_logs(patient_user_id)
    vitals_grouped = get_vitals_logs_grouped_by_day(patient_user_id)

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Patient Summary"
    ws_summary.append(["Field", "Value"])
    ws_summary.append(["Patient ID", patient["id"]])
    ws_summary.append(["Full Name", patient["full_name"]])
    ws_summary.append(["Email", patient["email"]])
    ws_summary.append(["DOB", patient["dob"] or ""])
    ws_summary.append(["Language", patient["language"] or ""])
    ws_summary.append(["Provider ID", patient["provider_id"] or ""])

    ws_glucose = wb.create_sheet("Glucose Logs")
    ws_glucose.append(
        [
            "ID", "Log Date", "Fasting BS", "Breakfast Notes", "Before Lunch BS",
            "Lunch Notes", "Before Dinner BS", "Dinner Notes", "Before Bedtime BS",
            "Nighttime Snack Notes", "Comments", "Entered By", "Created At",
        ]
    )
    for row in glucose_logs:
        ws_glucose.append(
            [
                row["id"], row["log_date"], row["fasting_bs"], row["breakfast_notes"],
                row["before_lunch_bs"], row["lunch_notes"], row["before_dinner_bs"],
                row["dinner_notes"], row["before_bedtime_bs"], row["nighttime_snack_notes"],
                row["comments"], row["entered_by_user_id"], row["created_at"],
            ]
        )

    ws_vitals = wb.create_sheet("BP Weight By Day")
    ws_vitals.append(
        ["Log Date", "AM BP", "Mid-Day BP", "PM BP", "Event BP", "AM Weight", "Mid-Day Weight", "PM Weight", "Event Weight", "Event Notes"]
    )
    for log_date, rows in vitals_grouped.items():
        day_map = {r["time_period"]: r for r in rows}
        am_row = day_map.get("AM")
        mid_row = day_map.get("Mid-Day")
        pm_row = day_map.get("PM")
        event_row = day_map.get("Event")
        event_notes = " | ".join([r["event_notes"] for r in rows if r["event_notes"]])

        ws_vitals.append(
            [
                log_date,
                am_row["blood_pressure"] if am_row else "",
                mid_row["blood_pressure"] if mid_row else "",
                pm_row["blood_pressure"] if pm_row else "",
                event_row["blood_pressure"] if event_row else "",
                am_row["weight_lbs"] if am_row else "",
                mid_row["weight_lbs"] if mid_row else "",
                pm_row["weight_lbs"] if pm_row else "",
                event_row["weight_lbs"] if event_row else "",
                event_notes,
            ]
        )

    for ws in wb.worksheets:
        for column_cells in ws.columns:
            length = max(len(str(cell.value or "")) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = min(max(length + 2, 12), 28)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def draw_wrapped_text(c, text, x, y, max_chars=90, line_height=11):
    text = str(text or "")
    if not text:
        return y
    words = text.split()
    line = ""
    lines = []
    for word in words:
        trial = f"{line} {word}".strip()
        if len(trial) <= max_chars:
            line = trial
        else:
            lines.append(line)
            line = word
    if line:
        lines.append(line)
    for item in lines:
        c.drawString(x, y, item)
        y -= line_height
    return y


def draw_box(c, x, y, w, h):
    c.rect(x, y, w, h)


def build_patient_pdf(patient_user_id: int) -> bytes:
    patient = get_user_by_id(patient_user_id)
    glucose_logs = get_glucose_logs(patient_user_id)
    vitals_grouped = get_vitals_logs_grouped_by_day(patient_user_id)
    chart_images = build_chart_images(patient_user_id)

    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter

    def header(title):
        c.setFont("Helvetica-Bold", 15)
        c.drawString(36, height - 32, title)
        c.setFont("Helvetica", 10)
        c.drawString(36, height - 50, f"Patient: {patient['full_name']}")
        c.drawString(280, height - 50, f"DOB: {patient['dob'] or ''}")
        c.drawString(36, height - 64, f"Email: {patient['email']}")
        c.drawString(280, height - 64, f"Language: {patient['language'] or ''}")
        c.drawString(36, height - 78, f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
        return height - 96

    y = header("Blood Sugar Log")
    c.setFont("Helvetica-Bold", 8)
    headers = ["Date", "Fasting\nBS", "Breakfast", "Before\nLunch BS", "Lunch", "Before\nDinner BS", "Dinner", "Before Bed /\nNight Snack"]
    col_widths = [58, 58, 64, 66, 62, 66, 66, 86]
    row_h = 28

    x = 36
    for i, text in enumerate(headers):
        draw_box(c, x, y - row_h, col_widths[i], row_h)
        tx = x + 3
        ty = y - 11
        for part in text.split("\n"):
            c.drawString(tx, ty, part)
            ty -= 9
        x += col_widths[i]

    c.setFont("Helvetica", 7)
    current_y = y - row_h
    for row in glucose_logs[:16]:
        current_y -= 30
        if current_y < 70:
            c.showPage()
            y = header("Blood Sugar Log")
            c.setFont("Helvetica-Bold", 8)
            x = 36
            for i, text in enumerate(headers):
                draw_box(c, x, y - row_h, col_widths[i], row_h)
                tx = x + 3
                ty = y - 11
                for part in text.split("\n"):
                    c.drawString(tx, ty, part)
                    ty -= 9
                x += col_widths[i]
            c.setFont("Helvetica", 7)
            current_y = y - row_h - 30

        values = [
            row["log_date"],
            row["fasting_bs"] or "",
            row["breakfast_notes"] or "",
            row["before_lunch_bs"] or "",
            row["lunch_notes"] or "",
            row["before_dinner_bs"] or "",
            row["dinner_notes"] or "",
            f'{row["before_bedtime_bs"] or ""} {row["nighttime_snack_notes"] or ""}'.strip(),
        ]
        x = 36
        for i, value in enumerate(values):
            draw_box(c, x, current_y, col_widths[i], 30)
            draw_wrapped_text(c, value, x + 2, current_y + 21, max_chars=max(10, int(col_widths[i] / 4.2)), line_height=8)
            x += col_widths[i]

    c.showPage()
    y = header("Blood Pressure and Weight Log")
    c.setFont("Helvetica-Bold", 8)
    bp_headers = ["Date", "AM", "Mid-Day", "PM", "Event", "Weight"]
    bp_widths = [78, 95, 95, 95, 95, 66]
    x = 36
    for i, text in enumerate(bp_headers):
        draw_box(c, x, y - 24, bp_widths[i], 24)
        c.drawString(x + 3, y - 14, text)
        x += bp_widths[i]

    c.setFont("Helvetica", 7)
    current_y = y - 24
    for log_date, rows in list(vitals_grouped.items())[:18]:
        current_y -= 30
        if current_y < 70:
            c.showPage()
            y = header("Blood Pressure and Weight Log")
            c.setFont("Helvetica-Bold", 8)
            x = 36
            for i, text in enumerate(bp_headers):
                draw_box(c, x, y - 24, bp_widths[i], 24)
                c.drawString(x + 3, y - 14, text)
                x += bp_widths[i]
            c.setFont("Helvetica", 7)
            current_y = y - 24 - 30

        day_map = {r["time_period"]: r for r in rows}
        am_row = day_map.get("AM")
        mid_row = day_map.get("Mid-Day")
        pm_row = day_map.get("PM")
        event_row = day_map.get("Event")

        am = am_row["blood_pressure"] if am_row else ""
        mid = mid_row["blood_pressure"] if mid_row else ""
        pm = pm_row["blood_pressure"] if pm_row else ""
        event = event_row["blood_pressure"] if event_row else ""
        weight_values = [str(r["weight_lbs"]) for r in rows if r["weight_lbs"] not in (None, "")]
        weight_text = ", ".join(weight_values)

        values = [log_date, am, mid, pm, event, weight_text]
        x = 36
        for i, value in enumerate(values):
            draw_box(c, x, current_y, bp_widths[i], 30)
            draw_wrapped_text(c, value, x + 2, current_y + 21, max_chars=max(10, int(bp_widths[i] / 4.5)), line_height=8)
            x += bp_widths[i]

        notes = " | ".join([f'{r["time_period"]}: {r["event_notes"]}' for r in rows if r["event_notes"]])
        if notes:
            current_y -= 14
            if current_y < 50:
                c.showPage()
                y = header("Blood Pressure and Weight Log")
                current_y = y - 40
                c.setFont("Helvetica", 7)
            c.drawString(42, current_y + 4, f"Notes: {notes}")

    for title, image_path in chart_images:
        c.showPage()
        y = header(title)
        c.drawImage(image_path, 36, 180, width=540, height=360, preserveAspectRatio=True, mask='auto')
        c.setFont("Helvetica", 9)
        if title == "Blood Sugar Trend":
            c.drawString(36, 150, "Reference lines shown at 70 and 180 mg/dL.")
        elif title == "Blood Pressure Trend":
            c.drawString(36, 150, "Reference lines shown at 130 systolic and 80 diastolic.")
        else:
            c.drawString(36, 150, "Use this chart to review weight direction over time.")

    c.save()
    buffer.seek(0)
    return buffer.getvalue()


def logout():
    st.session_state["user_id"] = None
    st.session_state["role"] = None
    st.rerun()


def show_auth():
    st.title(APP_TITLE)
    tab_login, tab_signup = st.tabs(["Log In", "Create Account"])

    with tab_login:
        with st.form("login_form"):
            email = st.text_input("Email")
            password = st.text_input("Password", type="password")
            submitted = st.form_submit_button("Log In")
        if submitted:
            user = get_user_by_email(email)
            if user and verify_password(password, user["password_hash"]):
                st.session_state["user_id"] = user["id"]
                st.session_state["role"] = user["role"]
                st.success("Logged in.")
                st.rerun()
            else:
                st.error("Invalid email or password.")

    with tab_signup:
        with st.form("signup_form"):
            full_name = st.text_input("Full Name")
            email = st.text_input("Email")
            password = st.text_input("Password", type="password")
            role = st.selectbox("Account Type", ["patient", "provider"])
            dob_value = st.date_input(
                "Date of Birth",
                min_value=date(1960, 1, 1),
                max_value=date(2026, 12, 31),
                value=date(1980, 1, 1),
            )
            language = st.selectbox("Language", ["English", "Spanish"])
            submitted = st.form_submit_button("Create Account")

        if submitted:
            if not full_name.strip():
                st.error("Full name is required.")
            elif not email.strip():
                st.error("Email is required.")
            elif len(password) < 6:
                st.error("Password must be at least 6 characters.")
            elif email_exists(email):
                st.error("That email already exists.")
            else:
                create_user(full_name, email, password, role, dob_value.isoformat(), language)
                st.success("Account created. Please log in.")


def glucose_form(prefix: str, default=None):
    default = default or {}
    raw_date = default.get("log_date")
    if isinstance(raw_date, str):
        raw_date = date.fromisoformat(raw_date)
    elif raw_date is None:
        raw_date = date.today()

    c1, c2 = st.columns(2)
    with c1:
        log_date = st.date_input("Log Date", value=raw_date, key=f"{prefix}_log_date")
        fasting_bs = st.number_input("Fasting Blood Sugar", min_value=0.0, step=1.0, value=float(default.get("fasting_bs") or 0.0), key=f"{prefix}_fasting_bs")
        before_lunch_bs = st.number_input("Before Lunch Blood Sugar", min_value=0.0, step=1.0, value=float(default.get("before_lunch_bs") or 0.0), key=f"{prefix}_before_lunch_bs")
        before_dinner_bs = st.number_input("Before Dinner Blood Sugar", min_value=0.0, step=1.0, value=float(default.get("before_dinner_bs") or 0.0), key=f"{prefix}_before_dinner_bs")
        before_bedtime_bs = st.number_input("Before Bedtime Blood Sugar", min_value=0.0, step=1.0, value=float(default.get("before_bedtime_bs") or 0.0), key=f"{prefix}_before_bedtime_bs")
    with c2:
        breakfast_notes = st.text_area("Breakfast Notes", value=default.get("breakfast_notes", "") or "", key=f"{prefix}_breakfast_notes")
        lunch_notes = st.text_area("Lunch Notes", value=default.get("lunch_notes", "") or "", key=f"{prefix}_lunch_notes")
        dinner_notes = st.text_area("Dinner Notes", value=default.get("dinner_notes", "") or "", key=f"{prefix}_dinner_notes")
        nighttime_snack_notes = st.text_area("Nighttime Snack Notes", value=default.get("nighttime_snack_notes", "") or "", key=f"{prefix}_nighttime_snack_notes")
        comments = st.text_area("Comments", value=default.get("comments", "") or "", key=f"{prefix}_comments")

    def normalize_num(x):
        return None if x in (0, 0.0) else float(x)

    return {
        "log_date": log_date.isoformat(),
        "fasting_bs": normalize_num(fasting_bs),
        "breakfast_notes": breakfast_notes.strip(),
        "before_lunch_bs": normalize_num(before_lunch_bs),
        "lunch_notes": lunch_notes.strip(),
        "before_dinner_bs": normalize_num(before_dinner_bs),
        "dinner_notes": dinner_notes.strip(),
        "before_bedtime_bs": normalize_num(before_bedtime_bs),
        "nighttime_snack_notes": nighttime_snack_notes.strip(),
        "comments": comments.strip(),
    }


def vitals_form(prefix: str, default=None):
    default = default or {}
    raw_date = default.get("log_date")
    if isinstance(raw_date, str):
        raw_date = date.fromisoformat(raw_date)
    elif raw_date is None:
        raw_date = date.today()

    log_date = st.date_input("Log Date", value=raw_date, key=f"{prefix}_v_log_date")
    options = ["AM", "Mid-Day", "PM", "Event"]
    current_period = default.get("time_period", "AM")
    index = options.index(current_period) if current_period in options else 0
    time_period = st.selectbox("Time Period", options, index=index, key=f"{prefix}_time_period")
    blood_pressure = st.text_input("Blood Pressure (example: 120/80)", value=default.get("blood_pressure", "") or "", key=f"{prefix}_blood_pressure")
    weight_lbs = st.number_input("Weight (lbs)", min_value=0.0, step=0.1, value=float(default.get("weight_lbs") or 0.0), key=f"{prefix}_weight_lbs")
    event_notes = st.text_area("Event Notes", value=default.get("event_notes", "") or "", key=f"{prefix}_event_notes")

    return {
        "log_date": log_date.isoformat(),
        "time_period": time_period,
        "blood_pressure": blood_pressure.strip(),
        "weight_lbs": None if weight_lbs in (0, 0.0) else float(weight_lbs),
        "event_notes": event_notes.strip(),
    }


def show_recent_summary(patient_user_id: int):
    recent_glucose = get_recent_glucose_logs(patient_user_id)
    grouped = get_vitals_logs_grouped_by_day(patient_user_id)

    col1, col2 = st.columns(2)
    with col1:
        st.markdown("**Recent Glucose Entries**")
        if not recent_glucose:
            st.caption("No glucose entries yet.")
        else:
            for row in recent_glucose:
                st.write(
                    f'{row["log_date"]}: fasting {row["fasting_bs"] or "-"}, '
                    f'before lunch {row["before_lunch_bs"] or "-"}, '
                    f'before dinner {row["before_dinner_bs"] or "-"}, '
                    f'bedtime {row["before_bedtime_bs"] or "-"}'
                )
                for label, value in [
                    ("Fasting", row["fasting_bs"]),
                    ("Before Lunch", row["before_lunch_bs"]),
                    ("Before Dinner", row["before_dinner_bs"]),
                    ("Before Bedtime", row["before_bedtime_bs"]),
                ]:
                    flag_text, level = glucose_flag(value)
                    if flag_text not in ("No reading", "In Range"):
                        render_flag(f"{label}: {flag_text}", level)
    with col2:
        st.markdown("**Recent BP / Weight Entries By Day**")
        if not grouped:
            st.caption("No vitals entries yet.")
        else:
            for log_date, rows in list(grouped.items())[:5]:
                day_map = {r["time_period"]: r for r in rows}
                pieces = []
                for period in ["AM", "Mid-Day", "PM", "Event"]:
                    row = day_map.get(period)
                    if row and (row["blood_pressure"] or row["weight_lbs"]):
                        pieces.append(f'{period}: BP {row["blood_pressure"] or "-"}, Wt {row["weight_lbs"] or "-"}')
                st.write(f"{log_date}: " + " | ".join(pieces))


def show_day_cards(patient_user_id: int):
    grouped_vitals = get_vitals_logs_grouped_by_day(patient_user_id)
    with st.expander("Blood Pressure / Weight Entries", expanded=True):
        if not grouped_vitals:
            st.info("No vitals entries yet.")
            return

        for log_date, rows in grouped_vitals.items():
            with st.container(border=True):
                st.markdown(f"### {log_date}")
                day_map = {r["time_period"]: r for r in rows}
                cols = st.columns(4)

                previous_weight = None
                for prior_row in sorted(rows, key=lambda r: ["AM", "Mid-Day", "PM", "Event"].index(r["time_period"]) if r["time_period"] in ["AM", "Mid-Day", "PM", "Event"] else 99):
                    if prior_row["weight_lbs"] not in (None, ""):
                        previous_weight = prior_row["weight_lbs"]
                        break

                for idx, period in enumerate(["AM", "Mid-Day", "PM", "Event"]):
                    row = day_map.get(period)
                    with cols[idx]:
                        st.markdown(f"**{period}**")
                        if row:
                            st.write(f'BP: {row["blood_pressure"] or "-"}')
                            st.write(f'Weight: {row["weight_lbs"] or "-"}')
                            bp_text, bp_level = bp_flag(row["blood_pressure"] or "")
                            if bp_text not in ("No reading", "In Range"):
                                render_flag(bp_text, bp_level)
                            wt_flag = weight_change_flag(row["weight_lbs"], previous_weight)
                            if wt_flag:
                                render_flag(wt_flag[0], wt_flag[1])
                            if row["event_notes"]:
                                st.caption(row["event_notes"])
                        else:
                            st.caption("No entry")

                st.divider()

                for row in rows:
                    label = f'{row["time_period"]} entry #{row["id"]}'
                    with st.expander(label):
                        with st.form(f"edit_vitals_{row['id']}"):
                            data = vitals_form(f"edit_vitals_{row['id']}", dict(row))
                            col1, col2 = st.columns(2)
                            save = col1.form_submit_button("Save Changes")
                            delete = col2.form_submit_button("Delete Entry")

                        if save:
                            if not validate_bp(data["blood_pressure"]):
                                st.error("Blood pressure must look like 120/80.")
                            else:
                                update_vitals_log(
                                    row["id"],
                                    patient_user_id,
                                    data["log_date"],
                                    data["time_period"],
                                    data["blood_pressure"],
                                    data["weight_lbs"],
                                    data["event_notes"],
                                )
                                st.success("Vitals entry updated.")
                                st.rerun()

                        if delete:
                            delete_vitals_log(row["id"], patient_user_id)
                            st.success("Vitals entry deleted.")
                            st.rerun()


def show_entry_management(patient_user_id: int):
    st.subheader("Manage Entries")

    glucose_logs = get_glucose_logs(patient_user_id)
    with st.expander("Glucose Entries", expanded=True):
        if not glucose_logs:
            st.info("No glucose entries yet.")
        for row in glucose_logs:
            label = f'{row["log_date"]} | Glucose Entry #{row["id"]}'
            with st.expander(label):
                with st.form(f"edit_glucose_{row['id']}"):
                    data = glucose_form(f"edit_glucose_{row['id']}", dict(row))
                    col1, col2 = st.columns(2)
                    save = col1.form_submit_button("Save Changes")
                    delete = col2.form_submit_button("Delete Entry")
                if save:
                    update_glucose_log(
                        row["id"],
                        patient_user_id,
                        data["log_date"],
                        data["fasting_bs"],
                        data["breakfast_notes"],
                        data["before_lunch_bs"],
                        data["lunch_notes"],
                        data["before_dinner_bs"],
                        data["dinner_notes"],
                        data["before_bedtime_bs"],
                        data["nighttime_snack_notes"],
                        data["comments"],
                    )
                    st.success("Glucose entry updated.")
                    st.rerun()
                if delete:
                    delete_glucose_log(row["id"], patient_user_id)
                    st.success("Glucose entry deleted.")
                    st.rerun()

    show_day_cards(patient_user_id)


def show_export_buttons(patient_user_id: int, label_prefix: str):
    patient = get_user_by_id(patient_user_id)
    pdf_bytes = build_patient_pdf(patient_user_id)
    excel_bytes = build_patient_excel(patient_user_id)

    col1, col2 = st.columns(2)
    with col1:
        st.download_button(
            label=f"{label_prefix} PDF Report + Charts",
            data=pdf_bytes,
            file_name=f"{patient['full_name'].replace(' ', '_').lower()}_clinic_report.pdf",
            mime="application/pdf",
        )
    with col2:
        st.download_button(
            label=f"{label_prefix} Excel File",
            data=excel_bytes,
            file_name=f"{patient['full_name'].replace(' ', '_').lower()}_trend_report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


def show_patient_dashboard(user):
    st.title(f"Welcome, {user['full_name']}")
    st.caption("Patient Dashboard")

    tab1, tab2, tab3, tab4, tab5 = st.tabs(
        ["Add Glucose Log", "Add BP / Weight", "Trends", "Manage My Entries", "Export"]
    )

    with tab1:
        with st.form("patient_glucose_form"):
            data = glucose_form("patient_glucose")
            submitted = st.form_submit_button("Save Glucose Entry")
        if submitted:
            add_glucose_log(
                user["id"], user["id"], data["log_date"], data["fasting_bs"], data["breakfast_notes"],
                data["before_lunch_bs"], data["lunch_notes"], data["before_dinner_bs"], data["dinner_notes"],
                data["before_bedtime_bs"], data["nighttime_snack_notes"], data["comments"],
            )
            st.success("Glucose entry saved.")
            for label, value in [("Fasting", data["fasting_bs"]), ("Before Lunch", data["before_lunch_bs"]), ("Before Dinner", data["before_dinner_bs"]), ("Before Bedtime", data["before_bedtime_bs"])]:
                flag_text, level = glucose_flag(value)
                if flag_text not in ("No reading", "In Range"):
                    render_flag(f"Saved with flag: {label} {flag_text}", level)

    with tab2:
        with st.form("patient_vitals_form"):
            data = vitals_form("patient_vitals")
            submitted = st.form_submit_button("Save BP / Weight Entry")
        if submitted:
            if not validate_bp(data["blood_pressure"]):
                st.error("Blood pressure must look like 120/80.")
            else:
                add_vitals_log(
                    user["id"], user["id"], data["log_date"], data["time_period"],
                    data["blood_pressure"], data["weight_lbs"], data["event_notes"],
                )
                st.success("Vitals entry saved.")
                bp_text, bp_level = bp_flag(data["blood_pressure"])
                if bp_text not in ("No reading", "In Range"):
                    render_flag(f"Saved with flag: {bp_text}", bp_level)

    with tab3:
        show_alert_summary(user["id"])
        st.divider()
        show_trend_charts(user["id"])

    with tab4:
        show_recent_summary(user["id"])
        st.divider()
        show_entry_management(user["id"])

    with tab5:
        st.write("Download a clinic-style printable PDF with trend charts or an Excel trend report.")
        show_export_buttons(user["id"], "Download")


def show_provider_dashboard(user):
    st.title(f"Welcome, {user['full_name']}")
    st.caption("Provider Dashboard")

    st.subheader("My Patients")
    search_text = st.text_input("Search patients by name or email", placeholder="Start typing a patient name or email")
    patients = get_provider_patients(user["id"], search_text)

    if patients:
        choices = {f'{r["full_name"]} ({r["email"]})': r["id"] for r in patients}
        selected_label = st.selectbox("Select patient", list(choices.keys()))
        selected_patient_id = choices[selected_label]
        selected_patient = get_user_by_id(selected_patient_id)
    else:
        selected_patient_id = None
        selected_patient = None
        st.info("No patients found for that search.")

    tab1, tab2, tab3, tab4, tab5 = st.tabs(
        ["Add Patient", "Assign Patient", "Delete Patient", "Patient Chart", "Export"]
    )

    with tab1:
        with st.form("provider_add_patient"):
            st.write("Create a new patient account")
            full_name = st.text_input("Patient Full Name")
            email = st.text_input("Patient Email")
            password = st.text_input("Temporary Password", type="password")
            dob_value = st.date_input(
                "Patient Date of Birth",
                min_value=date(1960, 1, 1),
                max_value=date(2026, 12, 31),
                value=date(1980, 1, 1),
                key="provider_add_patient_dob",
            )
            language = st.selectbox("Language", ["English", "Spanish"], key="provider_add_patient_language")
            submitted = st.form_submit_button("Create Patient")
        if submitted:
            if not full_name.strip() or not email.strip() or len(password) < 6:
                st.error("Enter name, email, and a password with at least 6 characters.")
            elif email_exists(email):
                st.error("A user with that email already exists.")
            else:
                create_user(full_name, email, password, "patient", dob_value.isoformat(), language)
                new_patient = get_user_by_email(email)
                assign_patient_to_provider(new_patient["id"], user["id"])
                st.success("Patient account created and assigned.")

    with tab2:
        unassigned = get_all_unassigned_patients()
        if not unassigned:
            st.info("No unassigned patient accounts.")
        else:
            options = {f'{r["full_name"]} ({r["email"]})': r["id"] for r in unassigned}
            selection = st.selectbox("Select patient account", list(options.keys()))
            if st.button("Assign Selected Patient"):
                assign_patient_to_provider(options[selection], user["id"])
                st.success("Patient assigned.")
                st.rerun()

    with tab3:
        if not selected_patient:
            st.info("Select a patient above to delete.")
        else:
            st.markdown(f"**Selected patient:** {selected_patient['full_name']}")
            confirm_name = st.text_input("Type the patient's full name to confirm")
            if st.button("Delete Patient Account"):
                if confirm_name.strip() != selected_patient["full_name"]:
                    st.error("Name did not match.")
                else:
                    delete_patient_account(selected_patient["id"], user["id"])
                    st.success("Patient account and related entries deleted.")
                    st.rerun()

    with tab4:
        if not selected_patient:
            st.info("Select a patient above to view the chart.")
        else:
            st.markdown(f"**Patient:** {selected_patient['full_name']}")
            st.markdown(f"**DOB:** {selected_patient['dob'] or 'Not set'}")
            st.markdown(f"**Language:** {selected_patient['language'] or 'English'}")

            show_alert_summary(selected_patient_id)
            st.divider()
            show_trend_charts(selected_patient_id)
            st.divider()
            show_entry_management(selected_patient_id)
            st.divider()

            st.subheader("Add Glucose Entry")
            with st.form("provider_glucose_entry"):
                data = glucose_form("provider_glucose")
                submitted = st.form_submit_button("Save Glucose Entry")
            if submitted:
                add_glucose_log(
                    selected_patient_id, user["id"], data["log_date"], data["fasting_bs"], data["breakfast_notes"],
                    data["before_lunch_bs"], data["lunch_notes"], data["before_dinner_bs"], data["dinner_notes"],
                    data["before_bedtime_bs"], data["nighttime_snack_notes"], data["comments"],
                )
                st.success("Glucose entry saved.")
                for label, value in [("Fasting", data["fasting_bs"]), ("Before Lunch", data["before_lunch_bs"]), ("Before Dinner", data["before_dinner_bs"]), ("Before Bedtime", data["before_bedtime_bs"])]:
                    flag_text, level = glucose_flag(value)
                    if flag_text not in ("No reading", "In Range"):
                        render_flag(f"Saved with flag: {label} {flag_text}", level)
                st.rerun()

            st.divider()
            st.subheader("Add Blood Pressure / Weight Entry")
            with st.form("provider_vitals_entry"):
                data = vitals_form("provider_vitals")
                submitted = st.form_submit_button("Save BP / Weight Entry")
            if submitted:
                if not validate_bp(data["blood_pressure"]):
                    st.error("Blood pressure must look like 120/80.")
                else:
                    add_vitals_log(
                        selected_patient_id, user["id"], data["log_date"], data["time_period"],
                        data["blood_pressure"], data["weight_lbs"], data["event_notes"],
                    )
                    st.success("Vitals entry saved.")
                    bp_text, bp_level = bp_flag(data["blood_pressure"])
                    if bp_text not in ("No reading", "In Range"):
                        render_flag(f"Saved with flag: {bp_text}", bp_level)
                    st.rerun()

    with tab5:
        if not selected_patient:
            st.info("Select a patient above to export.")
        else:
            st.write("Download a clinic-style printable PDF with trend charts or an Excel trend report.")
            show_export_buttons(selected_patient_id, "Download")


def main():
    st.set_page_config(page_title=APP_TITLE, layout="wide")
    init_db()

    if "user_id" not in st.session_state:
        st.session_state["user_id"] = None
        st.session_state["role"] = None

    with st.sidebar:
        st.header("Navigation")
        if st.session_state["user_id"]:
            current_user = get_user_by_id(st.session_state["user_id"])
            st.write(current_user["full_name"])
            st.write(current_user["role"].title())
            if st.button("Log Out"):
                logout()
        else:
            st.write("Please log in.")

    user_id = st.session_state["user_id"]
    if not user_id:
        show_auth()
        return

    user = get_user_by_id(user_id)
    if not user:
        logout()
        return

    if user["role"] == "provider":
        show_provider_dashboard(user)
    else:
        show_patient_dashboard(user)


if __name__ == "__main__":
    main()
